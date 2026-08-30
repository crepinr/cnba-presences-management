import argparse
import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from datetime import date, datetime

import config

try:
    import attendance_yaml_config
    attendance_yaml_config.bootstrap(config)
except Exception as yaml_config_error:
    # Si aucun YAML n'est configuré, on conserve le comportement historique.
    if "CNBA_CONFIG_YAML" in __import__("os").environ:
        raise


# ============================================================
# OUTILS DE NETTOYAGE
# ============================================================

def strip_accents(text: str) -> str:
    if text is None:
        return ""

    text = str(text)

    return "".join(
        character
        for character in unicodedata.normalize("NFD", text)
        if unicodedata.category(character) != "Mn"
    )


def clean_text(value) -> str:
    if value is None:
        return ""

    value = str(value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)

    return value.strip()


def normalize_name(value) -> str:
    return clean_text(value).upper()


def normalize_firstname(value) -> str:
    value = clean_text(value)

    if not value:
        return ""

    return value[:1].upper() + value[1:]


def normalize_sheet_name(value) -> str:
    value = clean_text(value).lower()
    value = strip_accents(value)

    return value


def normalize_status(value):
    """
    Nettoie les codes de présence.

    Exemples :
    - ' v ' devient 'V'
    - 'o' devient 'O'
    - cellule vide devient ''
    """
    if value is None:
        statut_original = ""
    else:
        statut_original = clean_text(value).upper()

    statut_clean = config.STATUS_MAPPING.get(
        statut_original,
        "inconnu"
    )

    return statut_original, statut_clean


def parse_birthdate(value):
    if value is None or value == "":
        return pd.NaT

    return pd.to_datetime(value, errors="coerce")


def create_swimmer_id(nom, prenom, date_naissance):
    """
    Crée un identifiant unique et stable pour chaque nageur.
    """
    date_txt = ""

    if pd.notna(date_naissance):
        date_txt = pd.to_datetime(date_naissance).strftime("%Y-%m-%d")

    return f"{normalize_name(nom)}_{normalize_firstname(prenom)}_{date_txt}"

def parse_training_date(value):
    """
    Convertit une date d'entraînement en vraie date pandas.

    Formats acceptés :
    - "01/09"
    - "1/9"
    - "01-09"
    - vraie date Excel lue par openpyxl
    - datetime Python
    - date Python
    - Timestamp pandas
    - texte complet : "2026-07-01" ou "2026-07-01 00:00:00"

    Important :
    Les formats courts jj/mm sont traités AVANT pd.to_datetime()
    pour éviter les interprétations ambiguës.
    """
    if value is None or value == "":
        return pd.NaT

    # ========================================================
    # 1. Cas où Excel/openpyxl renvoie déjà une vraie date
    # ========================================================

    if isinstance(value, pd.Timestamp):
        return pd.to_datetime(value).normalize()

    if isinstance(value, datetime):
        return pd.to_datetime(value.date())

    if isinstance(value, date):
        return pd.to_datetime(value)

    # ========================================================
    # 2. Cas texte
    # ========================================================

    value_text = clean_text(value)

    if not value_text:
        return pd.NaT

    # Format court belge : jj/mm ou jj-mm
    # Exemple : 01/07, 1/7, 01-07
    short_match = re.match(
        r"^(\d{1,2})[/-](\d{1,2})$",
        value_text
    )

    if short_match:
        day = int(short_match.group(1))
        month = int(short_match.group(2))

        year = (
            config.SEASON_START_YEAR
            if month >= 8
            else config.SEASON_END_YEAR
        )

        return pd.to_datetime(
            f"{year}-{month:02d}-{day:02d}",
            errors="coerce"
        )

    # Format complet ISO ou datetime converti en texte
    # Exemple : 2026-07-01 ou 2026-07-01 00:00:00
    full_date_match = re.match(
        r"^\d{4}-\d{1,2}-\d{1,2}(?:\s+\d{1,2}:\d{2}:\d{2})?$",
        value_text
    )

    if full_date_match:
        return pd.to_datetime(
            value_text,
            errors="coerce"
        ).normalize()

    # Format complet européen
    # Exemple : 01/07/2026 ou 1/7/2026
    european_full_match = re.match(
        r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$",
        value_text
    )

    if european_full_match:
        day = int(european_full_match.group(1))
        month = int(european_full_match.group(2))
        year = int(european_full_match.group(3))

        return pd.to_datetime(
            f"{year}-{month:02d}-{day:02d}",
            errors="coerce"
        )

    # ========================================================
    # 3. Cas numérique Excel éventuel
    # ========================================================
    # À garder en dernier, car les dates de ton fichier sont
    # normalement déjà converties par openpyxl.
    # ========================================================

    try:
        numeric_value = float(value_text)

        if numeric_value > 20000:
            return pd.to_datetime(
                numeric_value,
                unit="D",
                origin="1899-12-30",
                errors="coerce"
            ).normalize()

    except Exception:
        pass

    return pd.NaT

def detect_session_type(jour_code):
    """
    Déduit le type de séance à partir du code jour.

    Exemples :
    - mAM -> matin
    - jAM -> matin
    - l, m, j, v -> soir
    - s, d -> week-end
    """
    code = clean_text(jour_code)
    code_lower = code.lower()

    if any(keyword in code_lower for keyword in config.MORNING_KEYWORDS):
        return "matin"

    if code_lower in config.WEEKEND_CODES:
        return "week-end"

    if code_lower:
        return "soir"

    return "inconnu"

def normalize_for_matching(value) -> str:
    """
    Normalise un texte pour les comparaisons :
    - minuscules
    - sans accents
    - espaces nettoyés
    """
    value = clean_text(value).lower()
    value = strip_accents(value)
    return value


def is_summary_column_header(value) -> bool:
    """
    Détecte les colonnes de calcul / synthèse Excel à ignorer.

    Exemple :
    - Présences %
    - Présence %
    - Présences
    - %
    """
    header = normalize_for_matching(value)

    if not header:
        return False

    keywords = {
        normalize_for_matching(keyword)
        for keyword in config.SUMMARY_COLUMN_KEYWORDS
    }

    return any(keyword in header for keyword in keywords)


def get_last_session_column(ws) -> int:
    """
    Retourne la dernière colonne à considérer comme une séance.

    On lit la ligne DAY_CODE_ROW.
    Dès qu'on rencontre une colonne de synthèse comme "Présences %",
    on s'arrête juste avant.

    Si aucune colonne de synthèse n'est détectée, on garde ws.max_column.
    """
    for col in range(config.FIRST_SESSION_COLUMN, ws.max_column + 1):
        day_header = ws.cell(
            row=config.DAY_CODE_ROW,
            column=col
        ).value

        if is_summary_column_header(day_header):
            return col - 1

    return ws.max_column

# ============================================================
# EXPORT PICKLE
# ============================================================

def export_pickle_data(
    df_clean,
    df_nageurs,
    df_seances,
    audit_global,
    audit_dates,
    audit_statuts,
):
    """
    Exporte les dataframes en Pickle.

    Pickle est utilisé ici comme format de travail Python.
    Il accepte mieux les colonnes contenant encore des types mixtes
    après import brut depuis Excel.
    """
    config.PICKLE_DIR.mkdir(parents=True, exist_ok=True)

    df_clean.to_pickle(config.DATA_CLEAN_PICKLE)
    df_nageurs.to_pickle(config.NAGEURS_PICKLE)
    df_seances.to_pickle(config.SEANCES_PICKLE)

    audit_global.to_pickle(config.AUDIT_GLOBAL_PICKLE)
    audit_dates.to_pickle(config.AUDIT_DATES_PICKLE)
    audit_statuts.to_pickle(config.AUDIT_STATUTS_PICKLE)

    print(f"Fichiers Pickle générés dans : {config.PICKLE_DIR}")


# ============================================================
# ETAPE 1 — AUDIT DU FICHIER
# ============================================================

def audit_workbook(input_file):
    wb = load_workbook(input_file, data_only=True)

    audit_global = []

    for sheet_name in wb.sheetnames:
        audit_global.append({
            "sheet_name": sheet_name,
            "max_row": wb[sheet_name].max_row,
            "max_column": wb[sheet_name].max_column,
            "is_month_sheet": normalize_sheet_name(sheet_name) in [
                normalize_sheet_name(month)
                for month in config.MONTH_SHEETS
            ],
            "is_info_sheet": sheet_name in config.INFO_SHEETS,
        })

    return pd.DataFrame(audit_global)


def get_real_sheet_name(workbook, expected_sheet_name):
    """
    Retrouve le nom réel d'une feuille Excel indépendamment de la casse
    et des accents.

    Exemple :
    - config : juillet
    - Excel : Juillet
    """
    expected_normalized = normalize_sheet_name(expected_sheet_name)

    for sheet_name in workbook.sheetnames:
        if normalize_sheet_name(sheet_name) == expected_normalized:
            return sheet_name

    return None

def audit_month_sheets(input_file):
    wb = load_workbook(input_file, data_only=True)

    audit_dates = []
    audit_statuts = []

    for configured_sheet_name in config.MONTH_SHEETS:
        sheet_name = get_real_sheet_name(wb, configured_sheet_name)

        if sheet_name is None:
            audit_dates.append({
                "sheet_name": configured_sheet_name,
                "column": None,
                "raw_date": None,
                "parsed_date": pd.NaT,
                "jour_code": None,
                "expected_month": config.MONTH_NUMBER.get(
                    normalize_sheet_name(configured_sheet_name)
                ),
                "date_month": None,
                "date_mismatch": None,
                "issue": "Feuille mensuelle absente du fichier source",
            })
            continue

        ws = wb[sheet_name]

        last_session_column = get_last_session_column(ws)

        expected_month = config.MONTH_NUMBER.get(
            normalize_sheet_name(configured_sheet_name)
        )

        for col in range(config.FIRST_SESSION_COLUMN, last_session_column + 1):
            raw_date = ws.cell(
                row=config.DATE_ROW,
                column=col
            ).value

            jour_code = ws.cell(
                row=config.DAY_CODE_ROW,
                column=col
            ).value

            parsed_date = parse_training_date(raw_date)

            date_month = parsed_date.month if pd.notna(parsed_date) else None

            date_mismatch = (
                expected_month is not None
                and date_month is not None
                and expected_month != date_month
            )

            audit_dates.append({
                "sheet_name": sheet_name,
                "column": col,
                "raw_date": raw_date,
                "parsed_date": parsed_date,
                "jour_code": jour_code,
                "expected_month": expected_month,
                "date_month": date_month,
                "date_mismatch": date_mismatch,
                "issue": "",
            })

        for row in range(config.FIRST_SWIMMER_ROW, ws.max_row + 1):
            nom = ws.cell(
                row=row,
                column=config.NAME_COLUMN
            ).value

            prenom = ws.cell(
                row=row,
                column=config.FIRSTNAME_COLUMN
            ).value

            date_naissance = ws.cell(
                row=row,
                column=config.BIRTHDATE_COLUMN
            ).value

            if not nom or not prenom or not date_naissance:
                continue

            for col in range(config.FIRST_SESSION_COLUMN, last_session_column + 1):
                value = ws.cell(row=row, column=col).value

                statut_original, statut_clean = normalize_status(value)

                if statut_clean == "inconnu":
                    audit_statuts.append({
                        "sheet_name": sheet_name,
                        "row": row,
                        "column": col,
                        "nom": nom,
                        "prenom": prenom,
                        "value": value,
                        "statut_original": statut_original,
                        "statut_clean": statut_clean,
                    })

    return pd.DataFrame(audit_dates), pd.DataFrame(audit_statuts)

# ============================================================
# ETAPE 2 — IMPORT DES DONNEES
# ============================================================

def load_swimmers(input_file):
    """
    Importe la feuille Groupe.
    """
    df = pd.read_excel(
        input_file,
        sheet_name="Groupe",
        usecols="A:C",
        dtype=str
    )

    df.columns = [
        "nom",
        "prenom",
        "date_naissance",
    ]

    df["nom"] = df["nom"].apply(normalize_name)
    df["prenom"] = df["prenom"].apply(normalize_firstname)
    df["date_naissance"] = pd.to_datetime(
        df["date_naissance"],
        errors="coerce"
    )

    df = df.dropna(
        subset=[
            "nom",
            "prenom",
            "date_naissance",
        ]
    )

    df["nageur_id"] = df.apply(
        lambda row: create_swimmer_id(
            row["nom"],
            row["prenom"],
            row["date_naissance"]
        ),
        axis=1
    )

    df = df.drop_duplicates(subset=["nageur_id"])

    return df[
        [
            "nageur_id",
            "nom",
            "prenom",
            "date_naissance",
        ]
    ]

def parse_month_sheet(input_file, sheet_name):
    """
    Transforme une feuille mensuelle en format long.

    Les colonnes de synthèse Excel, comme 'Présences %',
    sont ignorées automatiquement.
    """
    wb = load_workbook(input_file, data_only=True)
    real_sheet_name = get_real_sheet_name(wb, sheet_name)

    if real_sheet_name is None:
        print(f"Feuille absente : {sheet_name}")
        return pd.DataFrame()

    ws = wb[real_sheet_name]

    last_session_column = get_last_session_column(ws)

    rows = []

    for row in range(config.FIRST_SWIMMER_ROW, ws.max_row + 1):
        nom_raw = ws.cell(
            row=row,
            column=config.NAME_COLUMN
        ).value

        prenom_raw = ws.cell(
            row=row,
            column=config.FIRSTNAME_COLUMN
        ).value

        birth_raw = ws.cell(
            row=row,
            column=config.BIRTHDATE_COLUMN
        ).value

        nom = normalize_name(nom_raw)
        prenom = normalize_firstname(prenom_raw)
        date_naissance = parse_birthdate(birth_raw)

        if not nom or not prenom or pd.isna(date_naissance):
            continue

        nageur_id = create_swimmer_id(
            nom,
            prenom,
            date_naissance
        )

        for col in range(config.FIRST_SESSION_COLUMN, last_session_column + 1):
            raw_date = ws.cell(
                row=config.DATE_ROW,
                column=col
            ).value

            jour_code = ws.cell(
                row=config.DAY_CODE_ROW,
                column=col
            ).value

            raw_status = ws.cell(
                row=row,
                column=col
            ).value

            date = parse_training_date(raw_date)

            if pd.isna(date):
                continue

            statut_original, statut_clean = normalize_status(raw_status)

            rows.append({
                "nageur_id": nageur_id,
                "nom": nom,
                "prenom": prenom,
                "date_naissance": date_naissance,
                "sheet_name": real_sheet_name,
                "mois": normalize_sheet_name(real_sheet_name),
                "date": date,
                "jour_code": clean_text(jour_code),
                "type_seance": detect_session_type(jour_code),
                "colonne_excel": col,
                "ligne_excel": row,
                "statut_original": statut_original,
                "statut_clean": statut_clean,
            })

    df = pd.DataFrame(rows)

    return df

def build_clean_dataset(input_file):
    all_months = []

    for sheet_name in config.MONTH_SHEETS:
        print(f"Import de la feuille : {sheet_name}")

        try:
            df_month = parse_month_sheet(input_file, sheet_name)
        except Exception as error:
            print(f"Erreur dans la feuille {sheet_name}: {error}")
            continue

        if not df_month.empty:
            all_months.append(df_month)

    if not all_months:
        raise ValueError("Aucune donnée mensuelle exploitable trouvée.")

    df = pd.concat(all_months, ignore_index=True)

    df = add_status_binary_columns(df)

    df["seance_comptabilisee"] = (
        df["statut_clean"]
        .isin(config.STATUTS_COMPTABILISES)
        .astype(int)
    )

    df["annee"] = df["date"].dt.year
    df["mois_numero"] = df["date"].dt.month
    df["semaine"] = df["date"].dt.isocalendar().week.astype(int)

    df["age_seance"] = (
        (df["date"] - df["date_naissance"]).dt.days / 365.25
    ).round(1)

    return df


def build_sessions_table(df_clean):
    """
    Crée une table unique des séances.
    """
    seances = (
        df_clean[
            [
                "sheet_name",
                "mois",
                "date",
                "jour_code",
                "type_seance",
                "colonne_excel",
            ]
        ]
        .drop_duplicates()
        .sort_values(["date", "colonne_excel"])
        .reset_index(drop=True)
    )

    seances["seance_id"] = [
        f"S{i + 1:03d}"
        for i in range(len(seances))
    ]

    return seances


def add_session_id(df_clean, seances):
    df = df_clean.merge(
        seances,
        on=[
            "sheet_name",
            "mois",
            "date",
            "jour_code",
            "type_seance",
            "colonne_excel",
        ],
        how="left"
    )

    return df


def apply_group_exclusions(df_clean):
    """
    Ajoute une colonne permettant de savoir si un nageur doit être exclu
    de l’analyse groupe.
    """
    df = df_clean.copy()

    df["exclude_from_group_analysis"] = (
        df["nageur_id"]
        .isin(config.EXCLUDED_SWIMMERS.keys())
    )

    df["exclusion_reason"] = (
        df["nageur_id"]
        .map(config.EXCLUDED_SWIMMERS)
        .fillna("")
    )

    return df

def get_expected_clean_statuses():
    """
    Récupère automatiquement les statuts propres depuis config.STATUS_MAPPING.
    """
    statuses = {
        clean_status
        for clean_status in config.STATUS_MAPPING.values()
        if clean_status is not None and str(clean_status).strip() != ""
    }

    return {
        str(status).strip()
        for status in statuses
    }


def add_status_binary_columns(df):
    """
    Crée automatiquement une colonne binaire pour chaque statut défini
    dans config.STATUS_MAPPING.

    Exemple :
    - present
    - absent
    - maladie
    - excuse
    - hors_analyse

    Si un nouveau statut est ajouté dans config.py, sa colonne est créée
    automatiquement.
    """
    df = df.copy()

    expected_statuses = get_expected_clean_statuses()

    for status in expected_statuses:
        df[status] = (
            df["statut_clean"] == status
        ).astype(int)

    df["statut_inconnu"] = (
        df["statut_clean"] == "inconnu"
    ).astype(int)

    return df

# ============================================================
# EXPORT EXCEL DE CONTROLE
# ============================================================

def export_clean_data(
    df_clean,
    df_nageurs,
    df_seances,
    audit_global,
    audit_dates,
    audit_statuts,
    output_file
):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_clean.to_excel(
            writer,
            sheet_name="data_clean",
            index=False
        )

        df_nageurs.to_excel(
            writer,
            sheet_name="nageurs",
            index=False
        )

        df_seances.to_excel(
            writer,
            sheet_name="seances",
            index=False
        )

        audit_global.to_excel(
            writer,
            sheet_name="audit_global",
            index=False
        )

        audit_dates.to_excel(
            writer,
            sheet_name="audit_dates",
            index=False
        )

        audit_statuts.to_excel(
            writer,
            sheet_name="audit_statuts",
            index=False
        )

    print(f"Fichier Excel de contrôle généré : {output_file}")


# ============================================================
# MAIN
# ============================================================

def main(group=None):
    if group is not None:
        config.set_active_group(group)

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Début audit du fichier...")
    audit_global = audit_workbook(config.INPUT_FILE)
    audit_dates, audit_statuts = audit_month_sheets(config.INPUT_FILE)

    print("Import des nageurs...")
    df_nageurs = load_swimmers(config.INPUT_FILE)

    print("Création du dataset importé...")
    df_clean = build_clean_dataset(config.INPUT_FILE)

    print("Création de la table des séances...")
    df_seances = build_sessions_table(df_clean)

    print("Ajout des identifiants de séance...")
    df_clean = add_session_id(df_clean, df_seances)

    print("Application des exclusions groupe...")
    df_clean = apply_group_exclusions(df_clean)

    print("Contrôles rapides...")
    print(f"Nombre de lignes importées : {len(df_clean)}")
    print(f"Nombre de nageurs : {df_clean['nageur_id'].nunique()}")
    print(f"Nombre de séances : {df_clean['seance_id'].nunique()}")

    if not audit_statuts.empty:
        print("\nAttention : certains statuts inconnus ont été détectés.")
        print(
            audit_statuts[
                [
                    "sheet_name",
                    "nom",
                    "prenom",
                    "value",
                ]
            ].head(20)
        )

    if not audit_dates.empty and "date_mismatch" in audit_dates.columns:
        date_mismatches = audit_dates[
            audit_dates["date_mismatch"] == True
        ]

        if not date_mismatches.empty:
            print("\nAttention : certaines dates ne correspondent pas au mois de la feuille.")
            print(
                date_mismatches[
                    [
                        "sheet_name",
                        "raw_date",
                        "parsed_date",
                        "expected_month",
                        "date_month",
                    ]
                ].head(20)
            )

    print("\nExport des fichiers Pickle...")

    export_pickle_data(
        df_clean=df_clean,
        df_nageurs=df_nageurs,
        df_seances=df_seances,
        audit_global=audit_global,
        audit_dates=audit_dates,
        audit_statuts=audit_statuts,
    )

    if config.EXPORT_EXCEL_CONTROL_FILE:
        print("\nExport du fichier Excel de contrôle...")

        export_clean_data(
            df_clean=df_clean,
            df_nageurs=df_nageurs,
            df_seances=df_seances,
            audit_global=audit_global,
            audit_dates=audit_dates,
            audit_statuts=audit_statuts,
            output_file=config.OUTPUT_FILE
        )

    print("Terminé.")


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--group",
        default=None,
        choices=config.available_groups(),
        help="Groupe à traiter.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    main(group=args.group)
